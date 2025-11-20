import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
import re
import numpy as np

# --- Configuration and Title ---
st.set_page_config(
    page_title="Stock Data Processor",
    layout="centered",
    initial_sidebar_state="auto"
)

st.title("📈 Streamlit Stock Analysis Flagging Tool")
st.markdown("Upload your Excel file to calculate industry averages, apply custom flags, sort, and download the formatted result.")

# Define the full combined flag names
NEW_SPECIAL_FLAG_NAME = '3B - 10B, PE>Ind, PE2 < PE1, EG2 > EG1'
NEW_SPECIAL_FLAG_NAME_2 = '10B+, PE < Ind, EG2 < EG1'

# --- File Uploader ---
uploaded_file = st.file_uploader(
    "1. Choose an Excel file (`.xlsx`)",
    type=['xlsx'],
    help="The file must contain columns like 'Industry', 'PE1', 'PE2', 'Market Cap in millions', 'EPS0', 'EPS1', and 'EPS2'. We will calculate EG1 and EG2 from EPS figures, format EG1/EG2 as percentages (0 decimals), PE1, PE2, PEG1, PEG2 to one decimal place, and EPS0, EPS1, EPS2 to two decimal places."
)

def find_robust_column(df_columns, required_key):
    """
    Finds the actual column name in the DataFrame based on a required key,
    handling casing, spacing, and punctuation variations (like parentheses).
    Returns the found column name or None.
    """
    # Function to clean a string: remove spaces, lowercase, remove punctuation (like parentheses)
    def clean_string(s):
        s = s.lower().replace(' ', '')
        # Remove common punctuation symbols for robustness
        s = re.sub(r'[()\[\]\{\}\.\-\_]', '', s)
        return s

    # 1. Standardize the required key for search
    standard_key = clean_string(required_key)
    
    # 2. Map cleaned column names to their original names
    standardized_columns = {clean_string(col): col for col in df_columns}
    
    # 3. Check for exact match of the standardized key
    if standard_key in standardized_columns:
        return standardized_columns[standard_key]

    # 4. Handle common abbreviations (e.g., matching 'marketcapmil')
    if 'marketcapmil' in standardized_columns:
        return standardized_columns['marketcapmil']

    # 5. Fallback fuzzy search 
    for std_col, original_col in standardized_columns.items():
        if standard_key in std_col or std_col in standard_key:
             return original_col
    
    return None

@st.cache_data
def process_excel_data(uploaded_file):
    """Reads, processes, calculates flags, sorts data, and returns a processed Pandas DataFrame."""
    try:
        df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error loading file: {e}")
        return None

    # --- 1. Data Preparation and Robust Column Mapping ---
    
    # Define the core columns needed for calculation (PE2 is now required for the flag logic)
    required_cols = {
        'industry': None,
        'pe1': None,
        'pe2': None, 
        'market cap (mil)': None, 
        'eps0': None,
        'eps1': None,
        'eps2': None,
    }
    
    # Map the required keys to the actual column names in the DataFrame
    for key in required_cols.keys():
        actual_col = find_robust_column(df.columns, key)
        if actual_col:
            required_cols[key] = actual_col
        else:
            # All these columns are now critical for calculation/flagging
            st.error(f"FATAL ERROR: Could not find a matching column for the required field: **'{key}'**.")
            st.info(f"The code now requires **'Industry', 'PE1', 'PE2', 'Market Cap (mil)', 'EPS0', 'EPS1', and 'EPS2'**. Please check your Excel headers: {list(df.columns)}")
            return None

    # --- Columns for Rounding ---
    
    # Identify columns for 1 decimal place rounding (PE1, PE2, PEG1, PEG2)
    one_decimal_cols_keys = ['pe1', 'pe2', 'peg1', 'peg2']
    actual_one_decimal_cols = []
    
    # Add mapped PE1 and PE2 first
    if required_cols['pe1']: actual_one_decimal_cols.append(required_cols['pe1'])
    if required_cols['pe2']: actual_one_decimal_cols.append(required_cols['pe2'])
    
    for key in one_decimal_cols_keys:
        actual_col = find_robust_column(df.columns, key)
        if actual_col:
            # Ensure PEG1 and PEG2 are included for rounding if they exist
            if actual_col not in actual_one_decimal_cols:
                actual_one_decimal_cols.append(actual_col)
            # Ensure PEG1, PEG2 are mapped if they weren't in the 7 core
            if key not in required_cols:
                 required_cols[key] = actual_col
            

    # Identify columns for 2 decimal place rounding (EPS0, EPS1, EPS2)
    two_decimal_cols_keys = ['eps0', 'eps1', 'eps2']
    actual_two_decimal_cols = [required_cols[k] for k in two_decimal_cols_keys if k in required_cols and required_cols[k]]
    
    
    # Create a simplified map for easy access
    col_map = {k: v for k, v in required_cols.items() if v is not None} # Only include mapped columns

    # All columns that need to be numeric: core calculation columns + rounding columns
    core_calc_cols = [col_map[k] for k in ['pe1', 'pe2', 'market cap (mil)', 'eps0', 'eps1', 'eps2']]
    core_calc_cols = [col for col in core_calc_cols if col is not None] # Filter out Nones
    
    # Combine all numeric columns and use set to ensure unique list of actual column names
    numeric_cols_actual = list(set(core_calc_cols + actual_one_decimal_cols + actual_two_decimal_cols)) 

    # --- Type Conversion and Initial Rounding ---
    for col in numeric_cols_actual:
        # Check if the column exists before converting
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Apply 1 decimal rounding to the PE/PEG columns
    for col in actual_one_decimal_cols:
        if col in df.columns:
            df[col] = df[col].round(1)
    
    # Apply 2 decimal rounding to the requested EPS columns
    for col in actual_two_decimal_cols:
        if col in df.columns:
            df[col] = df[col].round(2)
            
    # Drop rows where 'Industry' or 'PE1' is missing/invalid
    df.dropna(subset=[col_map['industry'], col_map['pe1']], inplace=True)
    
    # --- 2. Calculate Earnings Growth (EG1 and EG2) as DECIMAL VALUES ---
    
    # Calculate EG1: (EPS1 - EPS0) / EPS0 
    df['EG1'] = np.where(
        (df[col_map['eps0']] != 0) & df[col_map['eps0']].notna() & df[col_map['eps1']].notna(),
        (df[col_map['eps1']] - df[col_map['eps0']]) / df[col_map['eps0']],
        np.nan
    )
    
    # Calculate EG2: (EPS2 - EPS1) / EPS1
    df['EG2'] = np.where(
        (df[col_map['eps1']] != 0) & df[col_map['eps1']].notna() & df[col_map['eps2']].notna(),
        (df[col_map['eps2']] - df[col_map['eps1']]) / df[col_map['eps1']],
        np.nan
    )
    
    # Apply 2 decimal rounding to the raw decimal EG columns (e.g., 0.9394 becomes 0.94)
    df['EG1'] = df['EG1'].round(2)
    df['EG2'] = df['EG2'].round(2)
    
    # --- 3. Calculation of Industry Average PE1 ---
    industry_avg_pe = df.groupby(col_map['industry'])[col_map['pe1']].mean().reset_index()
    industry_avg_pe.rename(columns={col_map['pe1']: 'IndustryAvgPE1_Calc'}, inplace=True) # Use a temp name

    # Merge industry average back
    df = pd.merge(df, industry_avg_pe, on=col_map['industry'], how='left')

    # --- 4. Feature/Flag Creation ---
    
    # Round the calculated average to 1 decimal place (to match PE1, PE2, etc.)
    df['IndustryAvgPE1'] = df['IndustryAvgPE1_Calc'].round(1)
    df.drop(columns=['IndustryAvgPE1_Calc'], inplace=True) # Drop temp calculation column

    # Component Flag 1 (Existing logic): PE1 > Ind PE
    df['PE1 > Ind PE'] = (
        df[col_map['pe1']] > df['IndustryAvgPE1']
    )
    
    # Component Flag 2 (Existing logic): EG2 > EG1
    df['EG2_gt_EG1'] = df['EG2'] > df['EG1']

    # Component Flag 3 (Logic kept but column is removed from output): PE1 < Ind PE
    # We must calculate this boolean series for use in the combined flag below
    pe1_lt_ind_pe_series = (df[col_map['pe1']] < df['IndustryAvgPE1'])
    
    # Component Flag 4 (Logic kept but column is removed from output): EG2 < EG1
    # We must calculate this boolean series for use in the combined flag below
    eg2_lt_eg1_series = df['EG2'] < df['EG1']

    # COMBINED Flag 1 (Existing logic)
    df[NEW_SPECIAL_FLAG_NAME] = (
        df['PE1 > Ind PE'] & # 1. PE > Industry Avg
        df['EG2_gt_EG1'] &   # 2. EG2 > EG1 (Growth is accelerating)
        (df[col_map['pe2']] < df[col_map['pe1']]) &   # 3. PE2 < PE1 (Multiple is decreasing - calculated inline)
        (df[col_map['market cap (mil)']] >= 3000) & # 4. Market Cap >= 3B (3000M)
        (df[col_map['market cap (mil)']] <= 10000)  # 5. Market Cap <= 10B (10000M)
    )

    # COMBINED Flag 2 (New logic - uses the series calculated above)
    df[NEW_SPECIAL_FLAG_NAME_2] = (
        pe1_lt_ind_pe_series & # 1. PE < Industry Avg (using the series)
        eg2_lt_eg1_series &   # 2. EG2 < EG1 (Growth is decelerating - using the series)
        (df[col_map['market cap (mil)']] >= 10000) # 3. Market Cap >= 10B (10000M)
    )
    
    # --- 5. Clean up temporary and unwanted component flag columns ---
    # The flags 'PE1 < Ind PE' and 'EG2_lt_EG1' are now only used internally, so we don't need to add them.
    # We still need 'PE1 > Ind PE' and 'EG2_gt_EG1' for the first combined flag's logic, 
    # and they can remain in the output as they were not explicitly asked to be removed.
    
    # --- 6. Sorting ---
    df_sorted = df.sort_values(
        by=[col_map['industry'], col_map['pe1']],
        ascending=[True, False]
    )
    
    # Reset index for clean export
    df_sorted.reset_index(drop=True, inplace=True)
    
    # --- 7. Column Renaming and Ordering ---
    
    # Rename the core columns to standardized names for consistency and shading function compliance
    standardized_names = {
        col_map['industry']: 'Industry',
        col_map['pe1']: 'PE1',
        col_map['pe2']: 'PE2', 
        col_map['market cap (mil)']: 'Market Cap (mil)',
        col_map['eps0']: 'EPS0',
        col_map['eps1']: 'EPS1',
        col_map['eps2']: 'EPS2',
    }
    
    # Preserve original names for all other mapped columns (like PEG1, PEG2, etc.)
    core_keys = ['industry', 'pe1', 'pe2', 'market cap (mil)', 'eps0', 'eps1', 'eps2']
    for key, actual_col in col_map.items():
        if actual_col and key not in core_keys:
            standardized_names[actual_col] = actual_col # Keep original name

    df_sorted.rename(columns=standardized_names, inplace=True)
    
    # Define all calculated columns that should appear in the final output (6 total)
    calculated_cols = ['EG1', 'EG2', 'IndustryAvgPE1', 'PE1 > Ind PE', 'EG2_gt_EG1', NEW_SPECIAL_FLAG_NAME, NEW_SPECIAL_FLAG_NAME_2]
    
    # Get all columns currently in the DataFrame
    all_cols = list(df_sorted.columns)
    
    # Filter out the calculated columns from the non-calculated list (these are the 'original' columns)
    original_non_calculated_cols = [
        col for col in all_cols
        if col not in calculated_cols
    ]
    
    # --- Column Reordering Logic ---
    
    # 1. Find the index of 'EPS2' in the list of original columns
    try:
        eps2_index = original_non_calculated_cols.index('EPS2')
    except ValueError:
        eps2_index = -1 # Fallback if EPS2 is somehow not found

    # Calculated columns that should trail at the end 
    # Note: 'PE1 < Ind PE' and 'EG2_lt_EG1' are intentionally excluded here.
    trailing_calculated_cols = ['IndustryAvgPE1', 'PE1 > Ind PE', 'EG2_gt_EG1', NEW_SPECIAL_FLAG_NAME, NEW_SPECIAL_FLAG_NAME_2]
    
    if eps2_index != -1:
        # Split the list of original columns:
        before_eps2 = original_non_calculated_cols[:eps2_index + 1] 
        after_eps2 = original_non_calculated_cols[eps2_index + 1:]
        
        # Build the new final column list: (Before EPS2) + (EG1, EG2) + (After EPS2) + (Trailing Flags)
        final_output_cols = before_eps2 + ['EG1', 'EG2'] + after_eps2 + trailing_calculated_cols
    else:
        # Fallback to previous logic if EPS2 wasn't found
        final_output_cols = original_non_calculated_cols + calculated_cols
    
    # Select and reorder the DataFrame, ensuring we only select columns that exist
    df_final = df_sorted[[col for col in final_output_cols if col in df_sorted.columns]]
    
    return df_final

def apply_shading_and_save(df):
    """Applies openpyxl formatting to a dataframe and saves to BytesIO."""
    output = BytesIO()
    
    # Use Pandas to export to the buffer, which creates the structure
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Processed Stocks', index=False)
    
    # Load the workbook from the same buffer to apply styles
    wb = load_workbook(output)
    ws = wb['Processed Stocks']

    # --- Header Styling ---
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid") # Black fill
    header_font = Font(color="FFFFFF", bold=True) # White, bold font
    
    # Apply header style to the entire first row (column names)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # --- Row Shading and Number Formatting ---
    
    # Define the light gray fill for alternating rows
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Identify column indices for Industry and Percentage columns
    industry_col = None
    percentage_cols = ['EG1', 'EG2']
    percentage_col_indices = {}
    
    for col_num, cell in enumerate(ws[1], start=1):
        if cell.value == "Industry": 
            industry_col = col_num
        if cell.value in percentage_cols:
             percentage_col_indices[cell.value] = col_num

    # Apply alternating shading by industry
    if industry_col:
        current_industry = None
        gray_toggle = False
        # Start from row 2 (index 2) for data
        for row in range(2, ws.max_row + 1):
            industry_value = ws.cell(row=row, column=industry_col).value
            
            # Toggle shading when industry changes
            if industry_value != current_industry:
                if industry_value is not None:
                    gray_toggle = not gray_toggle
                current_industry = industry_value
            
            # Apply fill if toggled
            if gray_toggle:
                for col in range(1, ws.max_column + 1):
                    # Ensure we don't overwrite the header style
                    if row > 1:
                        ws.cell(row=row, column=col).fill = gray_fill
    
    # Apply percentage number format to EG1 and EG2 columns
    # The format "0%" displays the raw decimal value (e.g., 0.94) as 94%
    percentage_format = "0%"
    for col_name, col_num in percentage_col_indices.items():
        # Apply format to all data rows (starting from row 2)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            # Only apply format if the value is not None (i.e., not a NaN result from division by zero)
            if cell.value is not None:
                cell.number_format = percentage_format


    # Save the modified workbook back to a new BytesIO buffer
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# --- Main App Logic ---
if uploaded_file is not None:
    st.info("File successfully uploaded. Processing data...")
    
    df_processed = process_excel_data(uploaded_file)
    
    if df_processed is not None and not df_processed.empty:
        st.success(f"Processing complete! {len(df_processed)} records analyzed.")
        
        # Display Summary
        st.subheader("Processing Summary")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Records", len(df_processed))
        col2.metric(f"'{NEW_SPECIAL_FLAG_NAME}' (True)", df_processed[NEW_SPECIAL_FLAG_NAME].sum())
        col3.metric(f"'{NEW_SPECIAL_FLAG_NAME_2}' (True)", df_processed[NEW_SPECIAL_FLAG_NAME_2].sum())
        # The number of calculated columns is reduced from 8 to 6 in the final output
        col4.metric("New Calculated Columns", 6) 
        
        # Display the first few rows of the data
        st.subheader("Preview of Processed Data (First 10 Rows) - Check Column Order")
        # Displaying the raw decimal value (e.g., 0.94) in Streamlit is expected for this method.
        st.dataframe(df_processed.head(10), use_container_width=True)

        # --- Download Button ---
        st.markdown("---")
        st.subheader("2. Download Results")
        
        # Prepare the final excel file for download
        excel_bytes = apply_shading_and_save(df_processed)
        
        st.download_button(
            label="Download stocks_ranked_with_flags.xlsx",
            data=excel_bytes,
            file_name="stocks_ranked_with_flags.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        st.markdown(
            f"""
            <div style="font-size: 0.9em; color: gray;">
                The downloaded file is sorted by Industry (ASC) and PE1 (DESC) and includes alternating row shading.
                EG1 and EG2 are formatted as percentages (e.g., 0.94 displays as 94%) and placed next to EPS2.
                The new combined flags are: `{NEW_SPECIAL_FLAG_NAME}` and `{NEW_SPECIAL_FLAG_NAME_2}`.
            </div>
            """, unsafe_allow_html=True
        )
    elif df_processed is not None:
        st.warning("Processed DataFrame is empty. Check if your input file contains the required data and columns.")
